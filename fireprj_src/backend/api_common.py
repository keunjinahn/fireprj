print ("module [backend.api_common] loaded")

import hashlib
from flask import make_response, jsonify, request, json, send_from_directory, g
from flask_restless import ProcessingException
from flask_restful import reqparse
from datetime import datetime, timedelta
import os
import json
from functools import wraps
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Color, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend import app, login_manager
from backend_model.table_fireprj import *
from backend import manager
db = DBManager.db

manager.create_api(FireSensorTbl
                   , results_per_page=10000
                   , url_prefix='/api/v1'
                   , collection_name='sensor'
                   , methods=['GET', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True)

manager.create_api(FireRepeaterTbl
                   , results_per_page=10000
                   , url_prefix='/api/v1'
                   , collection_name='repeater'
                   , methods=['GET', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True)

manager.create_api(FireReceiverTbl
                   , results_per_page=10000
                   , url_prefix='/api/v1'
                   , collection_name='receiver'
                   , methods=['GET', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True)

def get_realtime_sensor(result=None, **kw):
    res = result['objects']
    all_logs = EventLogTbl.query.all()
    for sensor in res:
        sensor['regist_status'] = True #등록상태: 센서의 데이터가 존재하면 무조건 True
        last_sensor_log = list(filter(lambda x: x.sensor_id == sensor['sensor_id'], all_logs))
        if len(last_sensor_log) > 0: last_sensor_log = last_sensor_log[-1]
        else: sensor['action_status'] = sensor['network_status'] = sensor['battery_status'] = False; continue
        offset_time = datetime.now() - timedelta(minutes=1)
        sensor['receiver_type'] = last_sensor_log.receiver_type
        sensor['event_datetime'] = last_sensor_log.event_datetime
        sensor['action_status'] = sensor['network_status'] = sensor['battery_status'] \
        = True if last_sensor_log.event_datetime >= offset_time else False #1분 이내의 데이터가 있으면 True

manager.create_api(FireSensorTbl
                   , results_per_page=10000
                   , url_prefix='/api/v1'
                   , collection_name='sensor_event'
                   , methods=['GET', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True
                   , postprocessors={
                       'GET_MANY': [get_realtime_sensor]
                   })

manager.create_api(EventTbl
                   , results_per_page=10000
                   , url_prefix='/api/v1'
                   , collection_name='event_list'
                   , methods=['GET', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True)

manager.create_api(EventTbl
                   , results_per_page=10000
                   , url_prefix='/api/v1'
                   , collection_name='sensor_analysis'
                   , methods=['GET', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True)

def prePasswdUpdate(input_params=None, **kw):
    if 'user_pwd' in kw['data']:
        kw['data']['user_pwd'] = password_encoder_512(kw['data']['user_pwd'])
        
manager.create_api(UserTbl
                   , results_per_page=10000
                   , url_prefix='/api/v1'
                   , collection_name='user'
                   , methods=['GET', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True
                   , preprocessors={
                        'POST': [prePasswdUpdate],
                        'PATCH_SINGLE': [prePasswdUpdate],
                   })            

manager.create_api(CustomerTbl
                   , results_per_page=10000
                   , url_prefix='/api/v1'
                   , collection_name='customer'
                   , methods=['GET', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True)


@app.route('/make-data/event-list', methods=['GET'])
def write_event_list_api():
    file_path = 'C:\\Users\\ansieun\\project\\fireprj\\doc\\화재감시시스템 이벤트 리스트_20240215.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.get_sheet_by_name('Sheet1')
    
    data = []
    for i, row in enumerate(sheet.rows):
        row_data = []
        for j, col in enumerate(row):
            if j < 2: continue
            row_data.append(col.value)
        if row_data[0] is not None: data.append(row_data)
    del data[0]

    table_list = ['system_id_c', 'repeater_id_c', 'sensor_id_c', 'sensor_value_c', 'inout_id_c', 'event_desc']
    for i, row in enumerate(data):
        obj = EventTbl()
        setattr(obj, 'event_idx', i+1)
        for j, col in enumerate(row):
            setattr(obj, table_list[j], col)
        db.session.add(obj)
    db.session.commit()
    return make_response(jsonify(''), 200)

@app.route('/make-data/generating', methods=['POST'])
def generating_api():
    generate_option = json.loads(request.data)

    if 'receiver' in generate_option:
        receiver_len = generate_option['receiver']
        for i in range(receiver_len):
            receiver_id = i+1
            db.session.add(FireReceiverTbl(
                receiver_idx    = f'00001_001_{receiver_id:0>3}',
                fk_customer_idx = 1,                        #고객식별자: 1
                receiver_type   = 1,                        #수신기 타입: 1
                receiver_id     = receiver_id
            ))
        db.session.commit()

    if 'repeater' in generate_option:
        repeater_len = generate_option['repeater']
        receiver_data = FireReceiverTbl.query.all()
        for i, receiver in enumerate(receiver_data):
            for j in range(repeater_len):
                repeater_id = (i*repeater_len) + (j+1)
                db.session.add(FireRepeaterTbl(
                    repeater_idx    = f'00001_{receiver.receiver_id:0>3}_000_{repeater_id:0>3}',
                    fk_customer_idx = 1,                    #고객식별자: 1
                    receiver_id     = receiver.receiver_id,
                    system_id       = 0,                    #계통번호: 0
                    repeater_id     = repeater_id,
                ))
        db.session.commit()

    if 'sensor' in generate_option:
        sensor_len = generate_option['sensor']
        repeater_data = FireRepeaterTbl.query.all()
        for i, repeater in enumerate(repeater_data):
            for j in range(sensor_len):
                sensor_id = (i*sensor_len) + (j+1)
                db.session.add(FireSensorTbl(
                    sensor_idx      = repeater.repeater_idx + f'_{sensor_id:0>3}',
                    fk_customer_idx = 1,                    #고객식별자: 1
                    receiver_id     = repeater.receiver_id,
                    system_id       = 0,                    #계통번호: 0
                    repeater_id     = repeater.repeater_id,
                    sensor_id       = sensor_id,
                ))
        db.session.commit()

    return make_response(jsonify(''), 200)

@app.route('/api/v1/login', methods=['POST'])
def login_api():
    data = json.loads(request.data)
    result = ''
    print("111")
    if data['username'] is not None and data['password'] is not None:
        loginuser = db.session.query(UserTbl).filter(UserTbl.user_id == data["username"]).first()

        if loginuser is None:
            result = {'status': False, 'reason': 1}  # ID 없음
        else:
            if loginuser.user_pwd != password_encoder_512(data["password"]):
                result = {'status': False, 'reason': 2} # PW 틀림
                print("222")
            else:  # Login 성공
                if loginuser.user_status == 2:
                    result = {'status': False, 'reason': 3}  # Activation 안됨
                else:
                    print("333")
                    loginuser.token = generate_token(data['username'])
                    db.session.query(UserTbl).filter(UserTbl.user_id == data["username"])\
                        .update(dict(token=loginuser.token))
                    db.session.commit()

                    result = {'status': True, 'reason': 0, 'user': loginuser.serialize()}
    print("444:",result)
    return make_response(jsonify(result), 200)

@app.route("/api/v1/logout", methods=["POST"])
def logout_api():
    parser = reqparse.RequestParser()
    parser.add_argument("token", type=str, location="headers")
    token = parser.parse_args()["token"]
    result = ''
    if token is None:
        print("token is none")

    loginUser = UserTbl.query.filter_by(token=token).first()
    if loginUser is None:
        print("user is none")

    return make_response(jsonify(result), 200)

def generate_token(userID):
    m = hashlib.sha1()

    m.update(userID.encode('utf-8'))
    m.update(datetime.now().isoformat().encode('utf-8'))

    return m.hexdigest()

def check_token(search_params=None, **kw):
    parser = reqparse.RequestParser()
    parser.add_argument("token", type=str, location="headers")
    token = parser.parse_args()["token"]
    if token is None:
        raise ProcessingException(description="Not Authorized", code=410)
    user = UserTbl.query.filter_by(token=token).first()
    if user is None:
        raise ProcessingException(description="Not Authorized", code=411)

def check_token_single(search_params=None, **kw):
    parser = reqparse.RequestParser()
    parser.add_argument("token", type=str, location="headers")
    token = parser.parse_args()["token"]

    if token is None:
        raise ProcessingException(description="Not Authorized", code=410)

    user = UserTbl.query.filter_by(token=token).first()
    if user is None:
        raise ProcessingException(description="Not Authorized", code=411)

def password_encoder_512(password):
    h = hashlib.sha512()
    h.update(password.encode('utf-8'))
    return h.hexdigest()

@app.route('/api/v1/make_excel', methods=['POST'])
def make_excel_api():
    data = json.loads(request.data)
    page = data['page_name']
    db_router = {
        'sensor_manage': FireSensorTbl(),
        'repeater_manage': FireRepeaterTbl(),
        'receiver_manage': FireReceiverTbl(),
        # 'sensor_event': EventLogTbl(), #?
        'event_list': EventTbl(),
        # 'sensor_analysis': EventLogTbl(), #?
        'user': UserTbl(),
    }
    target_db = db_router[page]
    db_data = target_db.query.all()

    target_path = './download/'
    file_name = f'{page} {datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
    wb = Workbook()
    ws = wb.active

    ws.append(data['headers']) #헤더 삽입
    for row in db_data: #내용 삽입
        values = [getattr(row, col.name) for col in row.__table__.columns][1:]
        if 'manage' in page:
            values = [str(v).zfill(3) for v in values]
            values[1] = values[1].zfill(5)
        if page == 'event_list':
            values = [v if i == 0 or i == len(values) else str(v).zfill(3) for i, v in enumerate(values)]
        if page == 'user':
            del values[-1]
            values[1] = '****'
            values[3] = '사용' if values[3] == 1 else '미사용'
            values[4] = '관리' if values[3] == 1 else '조회'
        ws.append(values)

    bd_thin = Side(border_style='thin', color='000000')
    bd_all = Border(bd_thin, bd_thin, bd_thin, bd_thin)
    for row in ws.rows: #테두리, 정렬
        for cell in row:
            cell.border = bd_all
            cell.alignment = Alignment(horizontal='left')
    for cell in ws['1']: #헤더 스타일 지정
        cell.fill = PatternFill(fill_type='solid', fgColor=Color('a0a0a0'))
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(cell.col_idx)].width = 15 #열 크기 지정
    if page == 'sensor_manage': ws.column_dimensions['A'].width = 23
    if page == 'repeater_manage': ws.column_dimensions['A'].width = 17
    if page == 'receiver_manage': ws.column_dimensions['A'].width = 14
    if page == 'event_list': ws.column_dimensions['G'].width = 50

    wb.save(target_path+file_name)
    return make_response(jsonify({"filename":file_name}), 200)

@app.route('/api/v1/save_excel/<filename>', methods=['GET'])
def save_excel_api(filename):
    upload_path = os.getcwd() + '/download'
    return send_from_directory(upload_path, filename)

@app.route('/api/v1/sensor-log-chart', methods=['GET'])
def sensor_log_chart_api():
    sensor_id = request.args.get('sensor_id')
    data = EventLogTbl.query.filter_by(sensor_id=sensor_id)\
        .order_by(EventLogTbl.event_datetime.desc())\
        .offset(0).limit(10).all()
    values = []
    dates = []
    for d in data:
        values.append(float(d.sensor_value))
        dates.append(d.event_datetime)
    values = values[::-1]
    dates = dates[::-1]

    obj = {
        "objects": values,
        "dates": dates
    }
    return make_response(jsonify(obj), 200)