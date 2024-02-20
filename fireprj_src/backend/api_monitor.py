# -*- coding: utf-8 -*-
print ("module [backend.api_monitor] loaded")

from backend.api_common import *
from backend_model.table_typhoon import *
from sqlalchemy import sql
from openpyxl import load_workbook
from flask import send_from_directory, g
from backend.api_utils import *
from sqlalchemy import or_, and_, func
from datetime import datetime
import uuid
#damage_manage_download용 import
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
#school_info_update용 import
import sys
if sys.version_info[0] < 3:
    import urllib
else :
    import urllib.parse
from urllib.parse import urlparse
db = DBManager.db
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import random
import smtplib
global download_index
download_index = 0
# 대시보드 검색조건 데이터 조회
@app.route('/api/monitor/v1/dashboard/filter', methods=['GET'])
@token_required
def dashboard_filter_api():
    parser = reqparse.RequestParser()
    parser.add_argument("type", type=str, location='args', required=True)
    parser.add_argument("fk_user_id", type=str, location='args', required=True)
    parser.add_argument("h", type=str, location='args')
    args = parser.parse_args()
    fk_user_id = args["fk_user_id"]
    if args["type"] == "01" or args["type"] == "04":
        if args['h'] is None:
            years = set()
#             tps = db.session.query(Typhoons).filter(Typhoons.typhoon_number < '20210101').all()
            tps = db.session.query(Typhoons).all()
            for tp in tps:
                years.add(tp.typhoon_number[:4])

            years = sorted(years, reverse=True)

            data = [year for year in years]
            if args["type"] == "04" :
                data.insert(0,'전체')
        else:
            typhoons = set()

            tps = db.session.query(Typhoons.typhoon_name).filter(Typhoons.typhoon_number.like(args['h'] + "%")).filter(Typhoons.typhoon_v_type != 1).all()
            for tp in tps:
                typhoons.add(tp.typhoon_name)

            typhoons = sorted(typhoons)

            data = [typhoon for typhoon in typhoons]
    elif args["type"] == "02":
        typhoons = set()
        tps = db.session.query(Typhoons.typhoon_name).filter(Typhoons.typhoon_v_type == 2).filter(Typhoons.fk_user_id == fk_user_id).all()
        for tp in tps:
            typhoons.add(tp.typhoon_name)

        typhoons = sorted(typhoons)

        data = [typhoon for typhoon in typhoons]
    elif args["type"] == "03":
        if args['h'] is None:
            years = set()
            tps = db.session.query(Typhoons).filter(Typhoons.typhoon_v_type == 1).all()
            for tp in tps:
                years.add(tp.typhoon_number[:4])

            years = sorted(years, reverse=True)

            data = [year for year in years]
        else:
            typhoons = set()
            tps = db.session.query(Typhoons.typhoon_name).filter(Typhoons.typhoon_v_type == 1).filter(Typhoons.fk_user_id == g.user.id).all()
            for tp in tps:
                typhoons.add(tp.typhoon_name)

            typhoons = sorted(typhoons)

            data = [typhoon for typhoon in typhoons]

    return make_response(jsonify(data), 200)


# 대시보드 검색조건에 의한 데이터 리스트 조회
@app.route('/api/monitor/v1/dashboard/list', methods=['GET'])
@token_required
def dashboard_list_api():
    parser = reqparse.RequestParser()
    parser.add_argument("offset", type=int, location='args', default=0)
    parser.add_argument("limit", type=int, location='args', default=10)
    parser.add_argument("type", type=str, location='args', required=True)
    parser.add_argument("h1", type=str, location='args', required=True)
    parser.add_argument("h2", type=str, location='args')
    parser.add_argument("sort_by", type=str, location='args')
    parser.add_argument("sort_desc", type=str, location='args')
    args = parser.parse_args()

    if args['h2'] is not None:  # 태풍을 특정한 경우
        tp = db.session.query(Typhoons.id).filter(Typhoons.typhoon_name == args['h2']).first()
        qry_result = db.session.query(Accidents).filter(Accidents.fk_typhoon_id == tp.id)

        summary = db.session.query(Accidents.id, db.func.sum(Accidents.total_cost).label("TOTAL_COST"),
                                   db.func.count(Accidents.id).label("TOTAL_CN")).filter(Accidents.fk_typhoon_id == tp.id)
    else:
        if args['type'] == "01":  # 연도별
            qry_result = db.session.query(Accidents).filter(Accidents.acci_year == args['h1'])

            summary = db.session.query(Accidents.id, db.func.sum(Accidents.total_cost).label("TOTAL_COST"),
                                       db.func.count(Accidents.id).label("TOTAL_CN")).filter(Accidents.acci_year == args['h1'])
        elif args['type'] == "02":  # 방향별
            tp_ids = db.session.query(Typhoons.id).filter(Typhoons.typhoon_dir == args['h1']).all()

            qry_result = db.session.query(Accidents).filter(Accidents.fk_typhoon_id.in_(tp_ids))

            summary = db.session.query(Accidents.id, db.func.sum(Accidents.total_cost).label("TOTAL_COST"),
                                       db.func.count(Accidents.id).label("TOTAL_CN")).filter(Accidents.fk_typhoon_id.in_(tp_ids))
        elif args['type'] == "03":  # 세기별
            tp_ids = db.session.query(Typhoons.id).filter(Typhoons.typhoon_size == args['h1']).all()

            qry_result = db.session.query(Accidents).filter(Accidents.fk_typhoon_id.in_(tp_ids))

            summary = db.session.query(Accidents.id, db.func.sum(Accidents.total_cost).label("TOTAL_COST"),
                                       db.func.count(Accidents.id).label("TOTAL_CN")).filter(Accidents.fk_typhoon_id.in_(tp_ids))

    summary = summary.first()

    response = dict()
    response["total"] = qry_result.count()
    response["summary"] = dict(cost=str(summary.TOTAL_COST) if summary.TOTAL_COST is not None else "0",
                               count=str(summary.TOTAL_CN) if summary.TOTAL_CN is not None else "0")
    response["list"] = [dict(id=item.id,
                             school_name=item.school_name,
                             typhoon_id=item.fk_typhoon_id,
                             typhoon_name=item.typhoon_name,
                             total_cost=str(item.total_cost)) for item in qry_result]
    # 태풍 표기를 위해 페이지네이션 제거  조상래
    # .offset(args["offset"]).limit(args["limit"])

    return make_response(jsonify(response), 200)


# 대시보드 지도에서 태풍을 선택하여 피해학교 정보를 조회
# @app.route('/api/monitor/v1/dashboard/location', methods=['GET'])
# @token_required
# def dashboard_location_api():
#     parser = reqparse.RequestParser()
#     parser.add_argument("tp", type=str, location='args', required=True)
#     parser.add_argument("user_id", type=str, location='args', required=True)
#     args = parser.parse_args()
#     user = Users.query.filter_by(user_id=args['user_id']).first()
#     typhoon = Typhoons.query.filter_by(id=args['tp']).first()
#     if user is not None and typhoon is not None :
#         if typhoon.typhoon_v_type == 1 : #테스트 태풍이면
#             items = db.session.query(AccidentsPredict) \
#                 .filter(AccidentsPredict.latitude.isnot(None)) \
#                 .filter(AccidentsPredict.latitude < 42) \
#                 .filter(AccidentsPredict.latitude > 32) \
#                 .filter(AccidentsPredict.longitude.isnot(None)) \
#                 .filter(AccidentsPredict.longitude < 132) \
#                 .filter(AccidentsPredict.longitude > 123) \
#                 .filter(AccidentsPredict.fk_typhoon_id == args['tp']) \
#                 .filter(AccidentsPredict.fk_user_id == user.id) \
#                 .all()
#             response = [dict(id=item.id,
#                              school_id=item.fk_school_id,
#                              school_name=item.school_name,
#                              school_type=item.organ_gubun2_name,
#                              school_number=item.school.sub_school_id,
#                              mem_name=item.mem_name,
#                              latlng=[item.latitude, item.longitude],
#                              precipitation=item.precipitation,
#                              temperature=item.temperature,
#                              wind_speed=item.wind_speed,
#                              area=item.area.area_name,
#                              acci_number='',
#                              total_cost=item.total_payment_money) for item in items]
#         else :
#             items = db.session.query(Accidents) \
#                 .filter(Accidents.latitude.isnot(None)) \
#                 .filter(Accidents.latitude < 42) \
#                 .filter(Accidents.latitude > 32) \
#                 .filter(Accidents.longitude.isnot(None)) \
#                 .filter(Accidents.longitude < 132) \
#                 .filter(Accidents.longitude > 123) \
#                 .filter(Accidents.fk_typhoon_id == args['tp']) \
#                 .all()
#             response = [dict(id=item.id,
#                              school_id=item.sub_group_id,
#                              school_name=item.school_name,
#                              school_type=item.school_type,
#                              school_number=item.school_number,
#                              mem_name= item.subgroups.sub_group_name if item.subgroups is not None else '',
#                              latlng=[item.latitude, item.longitude],
#                              precipitation=item.precipitation,
#                              temperature=item.temperature,
#                              wind_speed=item.wind_speed,
#                              area=item.area.area_name,
#                              acci_number=item.acci_number,
#                              total_cost=item.total_cost) for item in items]
#
#     return make_response(jsonify(response), 200)

@app.route('/api/monitor/v1/dashboard/location', methods=['GET'])
@token_required
def dashboard_location_api():
    parser = reqparse.RequestParser()
    parser.add_argument("tp", type=str, location='args', required=True)
    parser.add_argument("fk_user_id", type=str, location='args', required=True)
    args = parser.parse_args()
    fk_user_id = args['fk_user_id']
    print("args :", args)
    typhoon = Typhoons.query.filter_by(id=args['tp']).first()
    response = []
    if typhoon is not None :
        if typhoon.typhoon_v_type == 1 : #테스트 태풍이면
            items = db.session.query(AccidentsPredict) \
                .filter(AccidentsPredict.latitude.isnot(None)) \
                .filter(AccidentsPredict.latitude < 42) \
                .filter(AccidentsPredict.latitude > 32) \
                .filter(AccidentsPredict.longitude.isnot(None)) \
                .filter(AccidentsPredict.longitude < 132) \
                .filter(AccidentsPredict.longitude > 123) \
                .filter(AccidentsPredict.fk_typhoon_id == args['tp']) \
                .filter(AccidentsPredict.fk_user_id == fk_user_id) \
                .all()
            print("items len :", len(items))
            for item in items:
                obj = dict(id=item.id,
                           school_id=item.sub_group_num,
                           school_name=item.school_name,
                           school_type=item.organ_gubun2_name,
                           school_number=item.sub_school_num,
                           mem_name=item.mem_name,
                           latlng=[item.latitude, item.longitude],
                           precipitation=item.precipitation,
                           temperature=item.temperature,
                           wind_speed=item.wind_speed,
                           area=item.area.area_name,
                           acci_number='',
                           total_cost=item.total_payment_money)
                response.append(obj)
        else :
            items = db.session.query(Accidents) \
                .filter(Accidents.latitude.isnot(None)) \
                .filter(Accidents.latitude < 42) \
                .filter(Accidents.latitude > 32) \
                .filter(Accidents.longitude.isnot(None)) \
                .filter(Accidents.longitude < 132) \
                .filter(Accidents.longitude > 123) \
                .filter(Accidents.fk_typhoon_id == args['tp']) \
                .all()

            for item in items:
                acci_contents = ''
                building_info = getJsonGroup12(item.s3_data_json)
                qurantee_info = get_school_info_from_json_ext(item,item.s3_data_json,'s3_2_group','s3_total_3','s3_2_group','s3_total_4')
                article = get_school_info_from_json_ext(item,item.s3_data_json,'s3_5_group','s3_total_5','s3_5_group','s3_total_7')
                acci_contents = '건물({0}건), 기본담보({1}건), 부속물/물품({2}건)'.format(building_info['count'],qurantee_info['count'],article['count'])
                obj = dict(id=item.id,
                   school_id=item.sub_group_id,
                   school_name=item.school_name,
                   school_type=item.school_type,
                   school_number=item.school_number,
                   mem_name= item.subgroups.sub_group_name if item.subgroups is not None else '',
                   latlng=[item.latitude, item.longitude],
                   precipitation=item.precipitation,
                   temperature=item.temperature,
                   wind_speed=item.wind_speed,
                   area=item.area.area_name,
                   acci_number=item.acci_number,
                   total_cost=item.total_cost,
                   acci_contents=acci_contents
                   )
                response.append(obj)
    return make_response(jsonify(response), 200)

# # 대시보드 검색조건에 의해 선택된 태풍 목록과 그에 대한 경로 조회
# @app.route('/api/monitor/v1/dashboard/paths', methods=['GET'])
# @token_required
# def dashboard_paths_api():
#     parser = reqparse.RequestParser()
#     parser.add_argument("type", type=str, location='args', required=True)
#     parser.add_argument("h1", type=str, location='args', required=True)
#     parser.add_argument("h2", type=str, location='args')
#     args = parser.parse_args()
#
#     response_list = []
#     tp_ids = set()
#
#     if args['h2'] is not None:  # 태풍을 특정한 경우
#         tp = db.session.query(Typhoons.id).filter(Typhoons.typhoon_name == args['h2']).first()
#         tp_ids.add(tp.id)
#     else:
#         if args['type'] == "01":  # 연도별
#             tps = db.session.query(Typhoons.id).filter(Typhoons.typhoon_number.like(args['h1'] + "%")).all()
#             for tp in tps:
#                 tp_ids.add(tp.id)
#         elif args['type'] == "02":  # 방향별
#             tps = db.session.query(Typhoons.id).filter(Typhoons.typhoon_dir == args['h1']).all()
#             for tp in tps:
#                 tp_ids.add(tp.id)
#         elif args['type'] == "03":  # 세기별
#             tps = db.session.query(Typhoons.id).filter(Typhoons.typhoon_size == args['h1']).all()
#             for tp in tps:
#                 tp_ids.add(tp.id)
#
#     for tp_id in tp_ids:
#         items = db.session.query(Paths) \
#             .filter(Paths.latitude.isnot(None)) \
#             .filter(Paths.latitude < 42) \
#             .filter(Paths.latitude > 32) \
#             .filter(Paths.longitude.isnot(None)) \
#             .filter(Paths.longitude < 132) \
#             .filter(Paths.longitude > 123) \
#             .filter(Paths.fk_typhoon_id == tp_id) \
#             .order_by(Paths.created_date.asc()) \
#             .all()
#
#         response = [dict(id=item.id,
#                          center_pressure=item.center_pressure,
#                          max_speed_s=item.max_speed_s,
#                          max_speed_h=item.max_speed_h,
#                          wind_radius=item.wind_radius,
#                          latlng=[item.latitude, item.longitude],
#                          storm_radius=item.storm_radius,
#                          direction=item.direction,
#                          move_speed_h=item.move_speed_h,
#                          modify_date=item.modify_date) for item in items]
#
#         tp = db.session.query(Typhoons).filter(Typhoons.id == tp_id).first()
#
#         data = {
#             "typhoon_id": tp.id,
#             "typhoon_name": tp.typhoon_name,
#             "typhoon_v_type": tp.typhoon_v_type,
#             "paths": response
#         }
#
#         response_list.append(data)
#
#     return make_response(jsonify(response_list), 200)

# 대시보드 검색조건에 의해 선택된 태풍 목록과 그에 대한 경로 조회
@app.route('/api/monitor/v1/dashboard/paths', methods=['GET'])
@token_required
def dashboard_paths_api():
    parser = reqparse.RequestParser()
    parser.add_argument("type", type=str, location='args', required=True)
    parser.add_argument("h1", type=str, location='args', required=True)
    parser.add_argument("h2", type=str, location='args')
    parser.add_argument("fk_user_id", type=str, location='args')
    args = parser.parse_args()
    fk_user_id = args['fk_user_id']
    response_list = []
    tp_ids = set()

    if args['h2'] is not None:  # 태풍을 특정한 경우
        tp = db.session.query(Typhoons.id).filter(Typhoons.typhoon_name == args['h2']).first()
        tp_ids.add(tp.id)
    else:
        if args['type'] == "01":  # 과거태풍
            tps = db.session.query(Typhoons.id).filter(Typhoons.typhoon_number.like(args['h1'] + "%")).filter(Typhoons.typhoon_v_type != 1).all()
            for tp in tps:
                tp_ids.add(tp.id)
        elif args['type'] == "02":  # 현재태풍(현재 사용안함)
            tps = db.session.query(Typhoons.id).filter(Typhoons.typhoon_dir == args['h1']).all()
            for tp in tps:
                tp_ids.add(tp.id)
        elif args['type'] == "03":  # 테스트태풍
            tps = db.session.query(Typhoons.id).filter(Typhoons.typhoon_number.like(args['h1'] + "%")).filter(Typhoons.typhoon_v_type == 1).filter(Typhoons.fk_user_id==fk_user_id).all()
            for tp in tps:
                tp_ids.add(tp.id)

    for tp_id in tp_ids:
        print("tp_id : ", tp_id)
#         items = db.session.query(Paths) \
#             .filter(Paths.latitude.isnot(None)) \
#             .filter(Paths.latitude < 42) \
#             .filter(Paths.latitude > 32) \
#             .filter(Paths.longitude.isnot(None)) \
#             .filter(Paths.longitude < 132) \
#             .filter(Paths.longitude > 123) \
#             .filter(Paths.fk_typhoon_id == tp_id) \
#             .order_by(Paths.created_date.asc()) \
#             .all()
        items = db.session.query(Paths) \
            .filter(Paths.latitude.isnot(None)) \
            .filter(Paths.latitude < 42) \
            .filter(Paths.latitude > 31.0) \
            .filter(Paths.longitude.isnot(None)) \
            .filter(Paths.longitude < 135) \
            .filter(Paths.longitude > 123.0) \
            .filter(Paths.fk_typhoon_id == tp_id) \
            .order_by(Paths.created_date.asc()) \
            .all()
        print("path items len : ", len(items))
        response = [dict(id=item.id,
                         center_pressure=item.center_pressure,
                         max_speed_s=item.max_speed_s,
                         max_speed_h=item.max_speed_h,
                         wind_radius=item.wind_radius,
                         latlng=[item.latitude, item.longitude],
                         storm_radius=item.storm_radius,
                         direction=item.direction,
                         move_speed_h=item.move_speed_h,
                         modify_date=item.modify_date) for item in items]

        tp = db.session.query(Typhoons).filter(Typhoons.id == tp_id).first()
        school_count = 0
        if tp.typhoon_v_type == 0 :
            school_count = Accidents.query.filter_by(fk_typhoon_id=tp.id).count()
        elif tp.typhoon_v_type == 1 :
            school_count = AccidentsPredict.query.filter_by(fk_typhoon_id=tp.id).count()
        data = {
            "typhoon_id": tp.id,
            "typhoon_name": tp.typhoon_name,
            "typhoon_v_type": tp.typhoon_v_type,
            "school_count": school_count,
            "paths": response
        }
#         print("data : ",data)
        response_list.append(data)

    return make_response(jsonify(response_list), 200)

# 통계화면에서 연도에 따른 통계 데이터 조회
@app.route('/api/monitor/v1/report/stats', methods=['GET'])
@token_required
def report_stats_api():
    parser = reqparse.RequestParser()
#     parser.add_argument("school_type", type=str, location='args')
#     parser.add_argument("area", type=int, location='args')
#
#     # 추가된 파라미터
    parser.add_argument("year_from", type=str, location='args')
    parser.add_argument("year_to", type=str, location='args')
    parser.add_argument("areas", type=str, location='args')
    parser.add_argument("typhoons", type=str, location='args')
    args = parser.parse_args()

    response = dict()
    print("args :",args)

#     year_num = int(args['year'])
#     school_type = args['school_type']
#     area = args['area']
    year_from = int(args['year_from'])
    year_to = int(args['year_to'])
    areas_string = str(args['areas']).split(',')[:-1]
    areas = [int(area) for area in areas_string]
    typhoons_string = str(args['typhoons']).split(',')[:-1]
    typhoons = [int(typhoon) for typhoon in typhoons_string]
    print("typhoons :",typhoons)
    if year_from is not None and year_to is not None:

        # 연도별
        q = TB_TYPOONREPORT.query.with_entities(
            db.func.substr(TB_TYPOONREPORT.sub2_sdate,1,4).label('sub2_sdate'),
            func.sum(TB_TYPOONREPORT.sub6_total_payment_money).label('cost'),
            func.count(TB_TYPOONREPORT.fk_typhoon_id).label('count')) \
            .filter(TB_TYPOONREPORT.fk_area_id.in_(areas)) \
            .filter(TB_TYPOONREPORT.fk_typhoon_id.in_(typhoons)) \
            .group_by(db.func.substr(TB_TYPOONREPORT.sub2_sdate,1,4)).all()

        response['by_year'] = [dict(year=item.sub2_sdate, cost=item.cost, count=item.count) for item in q]
        print("response['by_year'] :",response['by_year'] )
#
        # 지역별
        if areas is not None:
            q = TB_TYPOONREPORT.query.with_entities(
                TB_TYPOONREPORT.fk_area_id,
                Areas.area_name,
                func.sum(TB_TYPOONREPORT.sub6_total_payment_money).label('cost'),
                func.count(TB_TYPOONREPORT.fk_area_id).label('count')) \
                .join(Areas, Areas.id == TB_TYPOONREPORT.fk_area_id) \
                .filter(db.func.substr(TB_TYPOONREPORT.sub2_sdate,1,4) >= year_from) \
                .filter(db.func.substr(TB_TYPOONREPORT.sub2_sdate,1,4) <= year_to) \
                .filter(TB_TYPOONREPORT.fk_area_id.in_(areas)) \
                .filter(TB_TYPOONREPORT.fk_typhoon_id.in_(typhoons)) \
                .group_by(TB_TYPOONREPORT.fk_area_id).all()

            response['by_region'] = [dict(area=item.area_name, cost=item.cost, count=item.count) for item in q]
        else:
            response['by_region'] = []
        print("response['by_region'] :",response['by_region'] )
#         # 태풍별
        if typhoons is not None:
            q = TB_TYPOONREPORT.query.with_entities(
                TB_TYPOONREPORT.fk_typhoon_id,
                TB_TYPOONREPORT.sub2_disaster_naming,
                func.sum(TB_TYPOONREPORT.sub6_total_payment_money).label('cost'),
                func.count(TB_TYPOONREPORT.fk_typhoon_id).label('count')) \
                .join(Typhoons, Typhoons.id == TB_TYPOONREPORT.fk_typhoon_id) \
                .filter(db.func.substr(TB_TYPOONREPORT.sub2_sdate,1,4) >= year_from) \
                .filter(db.func.substr(TB_TYPOONREPORT.sub2_sdate,1,4) <= year_to) \
                .filter(TB_TYPOONREPORT.fk_area_id.in_(areas)) \
                .filter(TB_TYPOONREPORT.fk_typhoon_id.in_(typhoons)) \
                .group_by(TB_TYPOONREPORT.fk_typhoon_id).all()
            response['by_typhoons'] = []
            for item in q:
                typhoon_name = ''
                typhoon = Typhoons.query.filter_by(id=item.fk_typhoon_id).first()
                if typhoon is not None :
                    typhoon_name= typhoon.typhoon_name
                response['by_typhoons'].append(dict(typhoon=typhoon_name, cost=item.cost, count=item.count))
#             response['by_typhoons'] = [dict(typhoon=item.sub2_disaster_naming, cost=item.cost, count=item.count) for item in q]
        else:
            response['by_typhoons'] = []
        print("response['by_typhoons'] :",response['by_typhoons'] )
        # 설립별
        if typhoons is not None:
            print("year_from :",year_from)
            establish_list = TB_TYPOONREPORT.query.with_entities(
                TB_TYPOONREPORT.establish_gubun_name,
                func.sum(TB_TYPOONREPORT.sub6_total_payment_money).label('cost'),
                func.count(TB_TYPOONREPORT.fk_typhoon_id).label('count')) \
                .join(Typhoons, Typhoons.id == TB_TYPOONREPORT.fk_typhoon_id) \
                .filter(db.func.substr(TB_TYPOONREPORT.sub2_sdate,1,4) >= year_from) \
                .filter(db.func.substr(TB_TYPOONREPORT.sub2_sdate,1,4) <= year_to) \
                .filter(TB_TYPOONREPORT.fk_area_id.in_(areas)) \
                .filter(TB_TYPOONREPORT.fk_typhoon_id.in_(typhoons)) \
                .group_by(TB_TYPOONREPORT.establish_gubun_name).all()
            print("establish_list len  :",establish_list)
            response['by_establish'] = [dict(establish_gubun_name=item.establish_gubun_name, cost=item.cost, count=item.count) for item in establish_list]
        else:
            response['by_establish'] = []
#         print("response['by_establish'] :",response['by_establish'] )
        # 학교급별
        if typhoons is not None:

            school_type_list = TB_TYPOONREPORT.query.with_entities(
                TB_TYPOONREPORT.organ_gubun2_name,
                func.sum(TB_TYPOONREPORT.sub6_total_payment_money).label('cost'),
                func.count(TB_TYPOONREPORT.fk_typhoon_id).label('count')) \
                .join(Typhoons, Typhoons.id == TB_TYPOONREPORT.fk_typhoon_id) \
                .filter(db.func.substr(TB_TYPOONREPORT.sub2_sdate,1,4) >= year_from) \
                .filter(db.func.substr(TB_TYPOONREPORT.sub2_sdate,1,4) <= year_to) \
                .filter(TB_TYPOONREPORT.fk_area_id.in_(areas)) \
                .filter(TB_TYPOONREPORT.fk_typhoon_id.in_(typhoons)) \
                .group_by(TB_TYPOONREPORT.organ_gubun2_name).all()

            response['by_school_type'] = [dict(school_type=item.organ_gubun2_name, cost=item.cost, count=item.count) for item in school_type_list]
        else:
            response['by_school_type'] = []

    return make_response(jsonify(response), 200)


# @app.route('/api/monitor/v1/report/stats', methods=['GET'])
# @token_required
# def report_stats_api():
#     parser = reqparse.RequestParser()
# #     parser.add_argument("school_type", type=str, location='args')
# #     parser.add_argument("area", type=int, location='args')
# #
# #     # 추가된 파라미터
#     parser.add_argument("year_from", type=str, location='args')
#     parser.add_argument("year_to", type=str, location='args')
#     parser.add_argument("areas", type=str, location='args')
#     parser.add_argument("typhoons", type=str, location='args')
#     args = parser.parse_args()
#
#     response = dict()
#     print("args :",args)
#
# #     year_num = int(args['year'])
# #     school_type = args['school_type']
# #     area = args['area']
#     year_from = int(args['year_from'])
#     year_to = int(args['year_to'])
#     areas_string = str(args['areas']).split(',')[:-1]
#     areas = [int(area) for area in areas_string]
#     typhoons_string = str(args['typhoons']).split(',')[:-1]
#     typhoons = [int(typhoon) for typhoon in typhoons_string]
#     print("typhoons :",typhoons)
#     if year_from is not None and year_to is not None:
#
#         # 연도별
#         q = Accidents.query.with_entities(
#             Accidents.acci_year,
#             func.sum(Accidents.total_cost).label('cost'),
#             func.count(Accidents.fk_typhoon_id).label('count')) \
#             .filter(Accidents.acci_year >= year_from) \
#             .filter(Accidents.acci_year <= year_to) \
#             .filter(Accidents.fk_area_id.in_(areas)) \
#             .filter(Accidents.fk_typhoon_id.in_(typhoons)) \
#             .group_by(Accidents.acci_year).all()
#
#         response['by_year'] = [dict(year=item.acci_year, cost=item.cost, count=item.count) for item in q]
#         print("response['by_year'] :",response['by_year'] )
#         # 지역별
#         if areas is not None:
#             q = Accidents.query.with_entities(
#                 Accidents.fk_area_id,
#                 Areas.area_name,
#                 func.sum(Accidents.total_cost).label('cost'),
#                 func.count(Accidents.fk_area_id).label('count')) \
#                 .join(Areas, Areas.id == Accidents.fk_area_id) \
#                 .filter(Accidents.acci_year >= year_from) \
#                 .filter(Accidents.acci_year <= year_to) \
#                 .filter(Accidents.fk_area_id.in_(areas)) \
#                 .filter(Accidents.fk_typhoon_id.in_(typhoons)) \
#                 .group_by(Accidents.fk_area_id).all()
#
#             response['by_region'] = [dict(area=item.area_name, cost=item.cost, count=item.count) for item in q]
#         else:
#             response['by_region'] = []
# #         print("response['by_region'] :",response['by_region'] )
#         # 태풍별
#         if typhoons is not None:
#             q = Accidents.query.with_entities(
#                 Accidents.fk_typhoon_id,
#                 Typhoons.typhoon_name,
#                 func.sum(Accidents.total_cost).label('cost'),
#                 func.count(Accidents.fk_typhoon_id).label('count')) \
#                 .join(Typhoons, Typhoons.id == Accidents.fk_typhoon_id) \
#                 .filter(Accidents.acci_year >= year_from) \
#                 .filter(Accidents.acci_year <= year_to) \
#                 .filter(Accidents.fk_area_id.in_(areas)) \
#                 .filter(Accidents.fk_typhoon_id.in_(typhoons)) \
#                 .group_by(Accidents.fk_typhoon_id).all()
#
#             response['by_typhoons'] = [dict(typhoon=item.typhoon_name, cost=item.cost, count=item.count) for item in q]
#         else:
#             response['by_typhoons'] = []
# #         print("response['by_typhoons'] :",response['by_typhoons'] )
#         # 설립별
#         if typhoons is not None:
#             print("year_from :",year_from)
#             establish_list = Accidents.query.with_entities(
#                 Accidents.establish_gubun_name,
#                 func.sum(Accidents.total_cost).label('cost'),
#                 func.count(Accidents.fk_typhoon_id).label('count')) \
#                 .join(Typhoons, Typhoons.id == Accidents.fk_typhoon_id) \
#                 .filter(Accidents.acci_year >= year_from) \
#                 .filter(Accidents.acci_year <= year_to) \
#                 .filter(Accidents.fk_area_id.in_(areas)) \
#                 .filter(Accidents.fk_typhoon_id.in_(typhoons)) \
#                 .group_by(Accidents.establish_gubun_name).all()
#             print("establish_list len  :",establish_list)
#             response['by_establish'] = [dict(establish_gubun_name=item.establish_gubun_name, cost=item.cost, count=item.count) for item in establish_list]
#         else:
#             response['by_establish'] = []
# #         print("response['by_establish'] :",response['by_establish'] )
#         # 학교급별
#         if typhoons is not None:
#
#             school_type_list = Accidents.query.with_entities(
#                 Accidents.school_type,
#                 func.sum(Accidents.total_cost).label('cost'),
#                 func.count(Accidents.fk_typhoon_id).label('count')) \
#                 .join(Typhoons, Typhoons.id == Accidents.fk_typhoon_id) \
#                 .filter(Accidents.acci_year >= year_from) \
#                 .filter(Accidents.acci_year <= year_to) \
#                 .filter(Accidents.fk_area_id.in_(areas)) \
#                 .filter(Accidents.fk_typhoon_id.in_(typhoons)) \
#                 .group_by(Accidents.school_type).all()
#
#             response['by_school_type'] = [dict(school_type=item.school_type, cost=item.cost, count=item.count) for item in school_type_list]
#         else:
#             response['by_school_type'] = []
# #         print("response['by_school_type'] :",response['by_school_type'] )
# #
# #
# #         response['by_total_detail'] = {
# #             "name":'합계',
# #             "count_2016":0,
# #             "cost_2016":0,
# #             "count_2017":0,
# #             "cost_2017":0,
# #             "count_2018":0,
# #             "cost_2018":0,
# #             "count_2019":0,
# #             "cost_2019":0,
# #             "count_2020":0,
# #             "cost_2020":0,
# #             "count_sum":0,
# #             "cost_sum":0,
# #         }
# #
# #
# #         response['by_area_detail'] = []
# #         area_list = Areas.query.all()
# #         for area in area_list:
# #             response['by_area_detail'].append({
# #                 "area_name":area.area_name,
# #                 "count_2016":0,
# #                 "cost_2016":0,
# #                 "count_2017":0,
# #                 "cost_2017":0,
# #                 "count_2018":0,
# #                 "cost_2018":0,
# #                 "count_2019":0,
# #                 "cost_2019":0,
# #                 "count_2020":0,
# #                 "cost_2020":0,
# #                 "count_sum":0,
# #                 "cost_sum":0,
# #             })
# #
# #
# #         for year in response['by_year']:
# #             q = Accidents.query.with_entities(
# #                 Accidents.fk_area_id,
# #                 Areas.area_name,
# #                 func.sum(Accidents.total_cost).label('cost'),
# #                 func.count(Accidents.fk_area_id).label('count')) \
# #                 .join(Areas, Areas.id == Accidents.fk_area_id) \
# #                 .filter(Accidents.acci_year == year['year']) \
# #                 .filter(Accidents.fk_area_id.in_(areas)) \
# #                 .group_by(Accidents.fk_area_id).all()
# #             for item in q :
# #                 for _t in response['by_area_detail']:
# #                     if item.area_name == _t['area_name']:
# #                         _t['cost_' + year['year']] = item.cost
# #                         _t['count_' + year['year']] = item.count
# #                         _t['cost_sum'] += item.cost
# #                         _t['count_sum'] += item.count
# #                         response['by_total_detail']['cost_' + year['year']] += item.cost
# #                         response['by_total_detail']['count_' + year['year']] += item.count
# #                         response['by_total_detail']['cost_sum'] += item.cost
# #                         response['by_total_detail']['count_sum'] += item.count
# #                         break
# # #         print("response['by_area_detail'] :",response['by_area_detail'] )
# #
# #         #################typhoon
# #         response['by_typhoon_detail'] = []
# #         for t_id in typhoons:
# #             typhoon = Typhoons.query.filter_by(id=t_id).first()
# #             response['by_typhoon_detail'].append({
# #                 "typhoon_name":typhoon.typhoon_name,
# #                 "count_2016":0,
# #                 "cost_2016":0,
# #                 "count_2017":0,
# #                 "cost_2017":0,
# #                 "count_2018":0,
# #                 "cost_2018":0,
# #                 "count_2019":0,
# #                 "cost_2019":0,
# #                 "count_2020":0,
# #                 "cost_2020":0,
# #                 "count_sum":0,
# #                 "cost_sum":0,
# #             })
# #
# #
# #         for year in response['by_year']:
# #             q = Accidents.query.with_entities(
# #                 Accidents.fk_typhoon_id,
# #                 Typhoons.typhoon_name,
# #                 func.sum(Accidents.total_cost).label('cost'),
# #                 func.count(Accidents.fk_typhoon_id).label('count')) \
# #                 .join(Typhoons, Typhoons.id == Accidents.fk_typhoon_id) \
# #                 .filter(Accidents.acci_year == year['year']) \
# #                 .filter(Accidents.fk_typhoon_id.in_(typhoons)) \
# #                 .group_by(Accidents.fk_typhoon_id).all()
# #
# #             for item in q :
# #                 for _t in response['by_typhoon_detail']:
# #                     if item.typhoon_name == _t['typhoon_name']:
# #                         _t['cost_' + year['year']] = item.cost
# #                         _t['count_' + year['year']] = item.count
# #                         _t['cost_sum'] += item.cost
# #                         _t['count_sum'] += item.count
# #                         response['by_total_detail']['cost_' + year['year']] += item.cost
# #                         response['by_total_detail']['count_' + year['year']] += item.count
# #                         response['by_total_detail']['cost_sum'] += item.cost
# #                         response['by_total_detail']['count_sum'] += item.count
# #                         break
# #
# #         #################establish
# #         response['by_establish_detail'] = []
# #         establish_list = ['국립','공립','사립','BTL']
# #         for establish in establish_list:
# #             response['by_establish_detail'].append({
# #                 "establish_gubun_name":establish,
# #                 "count_2016":0,
# #                 "cost_2016":0,
# #                 "count_2017":0,
# #                 "cost_2017":0,
# #                 "count_2018":0,
# #                 "cost_2018":0,
# #                 "count_2019":0,
# #                 "cost_2019":0,
# #                 "count_2020":0,
# #                 "cost_2020":0,
# #                 "count_sum":0,
# #                 "cost_sum":0,
# #             })
# #
# #
# #         for year in response['by_year']:
# #             q = Accidents.query.with_entities(
# #                 Accidents.establish_gubun_name,
# #                 func.sum(Accidents.total_cost).label('cost'),
# #                 func.count(Accidents.fk_typhoon_id).label('count')) \
# #                 .join(Typhoons, Typhoons.id == Accidents.fk_typhoon_id) \
# #                 .filter(Accidents.acci_year == year['year']) \
# #                 .filter(Accidents.fk_typhoon_id.in_(typhoons)) \
# #                 .group_by(Accidents.establish_gubun_name).all()
# #
# #             for item in q :
# #                 for _t in response['by_establish_detail']:
# #                     if item.establish_gubun_name == _t['establish_gubun_name']:
# #                         _t['cost_' + year['year']] = item.cost
# #                         _t['count_' + year['year']] = item.count
# #                         _t['cost_sum'] += item.cost
# #                         _t['count_sum'] += item.count
# #                         response['by_total_detail']['cost_' + year['year']] += item.cost
# #                         response['by_total_detail']['count_' + year['year']] += item.count
# #                         response['by_total_detail']['cost_sum'] += item.cost
# #                         response['by_total_detail']['count_sum'] += item.count
# #                         break
# #
# #         #################학교급별
# #         response['by_scool_type_detail'] = []
# #         for school_type in school_type_list:
# #             response['by_scool_type_detail'].append({
# #                 "school_type":school_type.school_type,
# #                 "count_2016":0,
# #                 "cost_2016":0,
# #                 "count_2017":0,
# #                 "cost_2017":0,
# #                 "count_2018":0,
# #                 "cost_2018":0,
# #                 "count_2019":0,
# #                 "cost_2019":0,
# #                 "count_2020":0,
# #                 "cost_2020":0,
# #                 "count_sum":0,
# #                 "cost_sum":0,
# #             })
# #
# #         print("response['by_scool_type_detail'] :", response['by_scool_type_detail'])
# #         for year in response['by_year']:
# #             q = Accidents.query.with_entities(
# #                 Accidents.school_type,
# #                 func.sum(Accidents.total_cost).label('cost'),
# #                 func.count(Accidents.fk_typhoon_id).label('count')) \
# #                 .join(Typhoons, Typhoons.id == Accidents.fk_typhoon_id) \
# #                 .filter(Accidents.acci_year == year['year']) \
# #                 .filter(Accidents.fk_typhoon_id.in_(typhoons)) \
# #                 .group_by(Accidents.school_type).all()
# #             for item in q :
# #                 for _t in response['by_scool_type_detail']:
# #                     if item.school_type == _t['school_type']:
# #                         _t['cost_' + year['year']] = item.cost
# #                         _t['count_' + year['year']] = item.count
# #                         _t['cost_sum'] += item.cost
# #                         _t['count_sum'] += item.count
# #                         response['by_total_detail']['cost_' + year['year']] += item.cost
# #                         response['by_total_detail']['count_' + year['year']] += item.count
# #                         response['by_total_detail']['cost_sum'] += item.cost
# #                         response['by_total_detail']['count_sum'] += item.count
# #                         break
# #
#
# #
# #         print("response :",response['by_years_detail'] )
# #     for year in range(year_num-3, year_num+4):
# #         summary = db.session.query(Accidents.id, db.func.sum(Accidents.total_cost).label("TOTAL_COST"),
# #                                    db.func.count(Accidents.id).label("TOTAL_CN"))\
# #                                     .filter(Accidents.acci_year == str(year))
# #
# #         if school_type is not None:
# #             summary = summary.filter(Accidents.school_type == school_type)
# #
# #         if area is not None:
# #             summary = summary.filter(Accidents.fk_area_id == area)
# #
# #         summary = summary.first()
# #
# #         query = db.session.query(Typhoons).filter(Typhoons.typhoon_number.like(str(year) + "%"))
# #         tp_cnt = query.count()
# #
# #         # -3 ~ +3 7개년도 피해학교건수, 피해액수, 발생태풍건수
# #         response['summary_' + str(year)] = dict(acci_cost=str(summary.TOTAL_COST) if summary.TOTAL_COST is not None else "0",
# #                                    school_cnt=str(summary.TOTAL_CN) if summary.TOTAL_CN is not None else "0",
# #                                    tp_cnt=str(tp_cnt))
# #
# #
# #     # 태풍별 피해학교(막대), 태풍별 피해액수(막대)
# #     tp_data = list()
# #     query = db.session.query(Typhoons).filter(Typhoons.typhoon_number.like(str(year_num) + "%")).filter(Typhoons.typhoon_v_type != 1)
# #
# #     tps = query.all()
# #     for tp in tps:
# #         data = db.session.query(Accidents.id, db.func.sum(Accidents.total_cost).label("TOTAL_COST"),
# #                                    db.func.count(Accidents.id).label("TOTAL_CN")).filter(Accidents.fk_typhoon_id == tp.id)
# #
# #         if school_type is not None:
# #             data = data.filter(Accidents.school_type == school_type)
# #
# #         if area is not None:
# #             data = data.filter(Accidents.fk_area_id == area)
# #
# #         data = data.first()
# #
# #         tp_data.append(dict(tp_name=tp.typhoon_name,
# #                             acci_cost=str(data.TOTAL_COST) if data.TOTAL_COST is not None else "0",
# #                             school_cnt=str(data.TOTAL_CN) if data.TOTAL_CN is not None else "0"))
# #     response['tp_data'] = tp_data
#
#
#     # 지역별 피해학교(막대), 지역별 피해액수(막대)
# #     if area is not None:
# #         response['area_data'] = None
# #     else:
# #         region_data = list()
# #
# #         areas = db.session.query(Areas).all()
# #         for area in areas:
# #             data = db.session.query(Accidents.id, db.func.sum(Accidents.total_cost).label("TOTAL_COST"),
# #                                     db.func.count(Accidents.id).label("TOTAL_CN")).filter(Accidents.fk_area_id == area.id).first()
# #
# #             region_data.append(dict(area_name=area.area_name,
# #                                 acci_cost=str(data.TOTAL_COST) if data.TOTAL_COST is not None else "0",
# #                                 school_cnt=str(data.TOTAL_CN) if data.TOTAL_CN is not None else "0"))
# #         response['area_data'] = region_data
#
#     return make_response(jsonify(response), 200)

# 통계화면에서 연도에 따른 Top10 조회
@app.route('/api/monitor/v1/report/top10', methods=['GET'])
@token_required
def report_top10_api():
    parser = reqparse.RequestParser()
    parser.add_argument("year", type=str, location='args', required=True)
    parser.add_argument("school_type", type=str, location='args')
    parser.add_argument("area", type=int, location='args')
    args = parser.parse_args()

    school_type = args['school_type']
    area = args['area']

    # 피해학교 TOP10
    top10_schools = db.session.query(Accidents.id, Accidents.school_name, db.func.sum(Accidents.total_cost).label("TOTAL_COST"))\
                        .join(Typhoons, Typhoons.id == Accidents.fk_typhoon_id) \
                        .filter(Typhoons.typhoon_number.like(args['year'] + '%'))

    if school_type is not None:
        top10_schools = top10_schools.filter(Accidents.school_type == school_type)

    if area is not None:
        top10_schools = top10_schools.filter(Accidents.fk_area_id == area)

    top10_schools = top10_schools.group_by(Accidents.school_name) \
                        .order_by(Accidents.total_cost.desc()) \
                        .limit(10).all()

    # 피해지역 TOP10
    top10_areas = db.session.query(Accidents.id, Areas.area_name, db.func.sum(Accidents.total_cost).label("TOTAL_COST")) \
                        .join(Areas, Areas.id == Accidents.fk_area_id) \
                        .filter(Typhoons.typhoon_number.like(args['year'] + '%')) \

    if school_type is not None:
        top10_areas = top10_areas.filter(Accidents.school_type == school_type)

    if area is not None:
        top10_areas = top10_areas.filter(Accidents.fk_area_id == area)

    top10_areas = top10_areas.group_by(Accidents.fk_area_id) \
                        .order_by(sql.text('TOTAL_COST desc')) \
                        .limit(10).all()

    response = dict()
    response['top10_schools'] = top10_schools
    response['top10_areas'] = top10_areas

    return make_response(jsonify(response), 200)


@app.route('/api/monitor/v1/upload_excel', methods=['POST'])
@token_required
def upload_excel_api():
    check_token()
    file = request.files['file']

    newDir = app.config['UPLOAD_FOLDER'] + "/"
    try:
        os.mkdir(newDir)
    except:
        pass

    newFullPath = newDir + file.filename

    file.save(newFullPath)
    print("newFullPath :", newFullPath)
    wb = load_workbook(newFullPath, data_only=True)
    sheet = wb.active

    idx = 0
    #print("Accident add start")
    for row in sheet.rows:
        idx += 1

        if idx <= 2:
            continue
        #print("school_name :",row[9].value)
        acc = Accidents()
        if row[1].value is not None :
            pre_acc = Accidents.query.filter_by(school_number=row[11].value).first()
            if pre_acc is None :
                acc.acci_number = row[1].value
                acc.acci_year = row[5].value
                acc.school_id = row[8].value
                acc.school_name = row[9].value
                acc.school_number = row[11].value
                acc.school_type = row[14].value
                acc.school_address = row[16].value
                acc.latitude = row[17].value
                acc.longitude = row[18].value
                acc.acci_occu_date = row[19].value
                # acc.damage_desc = item['']
                acc.precipitation = row[24].value
                acc.temperature = row[26].value
                acc.wind_speed = row[27].value
                acc.typhoon_name = row[34].value
                acc.total_cost = row[129].value

                if acc.acci_occu_date is not None:
                    tp = db.session.query(Typhoons).filter(Typhoons.typhoon_start_date <= acc.acci_occu_date) \
                        .filter(Typhoons.typhoon_end_date >= acc.acci_occu_date).first()
                    if tp is not None:
                        acc.fk_typhoon_id = tp.id
                    else:
                        acc.fk_typhoon_id = -1
                else:
                    acc.fk_typhoon_id = -1

                sc = db.session.query(Schools).filter(Schools.school_number == acc.school_number).first()
                if sc is not None:
                    acc.fk_school_id = sc.id
                else:
                    acc.fk_school_id = -1

                if row[2].value is not None:
                    ar = db.session.query(Areas).filter(Areas.area_name.like("%" + row[2].value + "%")).first()
                    if ar is not None:
                        acc.fk_area_id = ar.id
                    else:
                        acc.fk_area_id = -1
                else:
                    acc.fk_area_id = -1

                DBManager.db.session.add(acc)

    DBManager.db.session.commit()

    return make_response(jsonify(''), 200)

'''
@app.route('/api/monitor/v1/upload_excel', methods=['POST'])
@token_required
def upload_excel_api():
    data = json.loads(request.data)
    data_list = json.loads(data['obj'])

    for item in data_list:
        acc = Accidents()
        acc.acci_number = item['재난번호'] if '재난번호' in item else ''
        acc.acci_year = item['사고년도'] if '사고년도' in item else ''
        acc.school_id = item['회원번호'] if '회원번호' in item else ''
        acc.school_name = item['학교명'] if '학교명' in item else ''
        acc.school_number = item['학교번호'] if '학교번호' in item else ''
        acc.school_type = item['기관종류'] if '기관종류' in item else ''
        acc.school_address = item['행정구역'] if '행정구역' in item else ''
        acc.latitude = item['위도'] if '위도' in item else 0
        acc.longitude = item['경도'] if '경도' in item else 0
        acc.acci_occu_date = item['발생일시'] if '발생일시' in item else '1970-01-01'
        #acc.damage_desc = item['']
        acc.precipitation = item['강수량'] if '강수량' in item else 0
        acc.temperature = item['온도'] if '온도' in item else 0
        acc.wind_speed = item['풍속'] if '풍속' in item else 0
        acc.typhoon_name = item['대형재난'] if '대형재난' in item else ''
        acc.total_cost = item['총지급액'] if '총지급액' in item else 0

        tp = db.session.query(Typhoons).filter(Typhoons.typhoon_start_date <= acc.acci_occu_date)\
            .filter(Typhoons.typhoon_end_date >= acc.acci_occu_date).first()
        if tp is not None:
            acc.fk_typhoon_id = tp.id
        else:
            acc.fk_typhoon_id = -1

        sc = db.session.query(Schools).filter(Schools.school_number == acc.school_number).first()
        if sc is not None:
            acc.fk_school_id = sc.id
        else:
            acc.fk_school_id = -1

        if '접수지역' in item:
            ar = db.session.query(Areas).filter(Areas.area_name.like("%" + item['접수지역'] + "%")).first()
            if ar is not None:
                acc.fk_area_id = ar.id
            else:
                acc.fk_area_id = -1
        else:
            acc.fk_area_id = -1

        DBManager.db.session.add(acc)

    DBManager.db.session.commit()

    return make_response(jsonify('Success'), 200)
'''


@app.route('/api/monitor/v1/attachment', methods=['POST'])
def set_attachment_images():
    print("1")
    check_token()
    print("2")
    print(request.files)
    if 'file' not in request.files:
        return make_response(jsonify({'result': False}), 400)
    print("3")
    f = request.files['file']
    if f.filename == '':
        return make_response(jsonify({'result': False}), 400)
    print("f.filename ",f.filename)
    upload_path = os.getcwd() + '/attachments/'
    print("upload_path :",upload_path)
    if f and '.' in f.filename:
        ext = f.filename.rsplit('.', 1)[1].lower()

        if ext in {'png', 'jpg', 'jpeg', 'gif'}:
            filename = str(uuid.uuid1()) + '.' + ext
            f.save(os.path.join(upload_path, filename))
            return make_response(jsonify({'result': True, 'filename': filename,'src_filename':f.filename}), 200)

@app.route('/api/monitor/v1/upload_file', methods=['POST'])
def upload_file_api():
    print("1")
    check_token()
    print("2")
    print(request.files)
    if 'file' not in request.files:
        return make_response(jsonify({'result': False}), 400)
    print("3")
    f = request.files['file']
    print("f.filename ",f.filename)
    if f.filename == '':
        return make_response(jsonify({'result': False}), 400)
    print("4")
    upload_path = os.getcwd() + '/uploads/'
    print("upload_path :",upload_path)
    if f and '.' in f.filename:
        filename = f.filename
        f.save(os.path.join(upload_path, filename))
        return make_response(jsonify({'result': True, 'filename': filename,'src_filename':f.filename}), 200)

@app.route('/api/monitor/v1/upload_file/<filename>', methods=['GET'])
def get_upload_fil_api(filename):
    upload_path = os.getcwd() + '/uploads/'
    return send_from_directory(upload_path, filename)

@app.route('/api/monitor/v1/attachment/<filename>', methods=['GET'])
def get_attachment_images(filename):
    upload_path = os.getcwd() + '/attachments/'
    return send_from_directory(upload_path, filename)

@app.route('/api/monitor/v1/download/<filename>', methods=['GET'])
def get_download_excel(filename):
    upload_path = os.getcwd() + '/download/excel/'
    return send_from_directory(upload_path, filename)

@app.route('/api/monitor/v1/update_damage_imagepath', methods=['POST'])
def update_damage_imagepath_api():
    data = json.loads(request.data)
    acci_past_id = data['acci_past_id']
#     print("data : ", data)
    acci_past = TB_TYPOONREPORT.query.filter_by(id=acci_past_id).first()
    if acci_past is not None :
        obj = {
            "image_file_path1":data['image_file_path1']
            ,"image_file_path2":data['image_file_path2']
            ,"image_file_path3":data['image_file_path3']
            ,"image_file_path4":data['image_file_path4']
            ,"image_file_path5":data['image_file_path5']
            ,"image_file_path6":data['image_file_path6']
        }
        db.session.query(TB_TYPOONREPORT).filter_by(id=acci_past_id).update(obj)
        db.session.commit()
        print("obj : ", obj)
    return make_response(jsonify({}), 200)

@app.route('/api/monitor/v1/data/<filename>', methods=['GET'])
def get_data_files(filename):
    upload_path = os.getcwd() + '/data/'
    return send_from_directory(upload_path, filename)

@app.route('/api/monitor/v1/acci_info_from_report', methods=['POST'])
def acci_info_from_report_api():
    Accidents.query.delete()
    db.session.commit()
    typhoon_report_list = TB_TYPOONREPORT.query.all()

    #print("Accident add start")
    for idx, typhoon_report in enumerate(typhoon_report_list):
        acc = Accidents()
        print("school name :",typhoon_report.school_name ,",idx :",idx + 1)
        acc.acci_number = typhoon_report.disaster_num
        acc.acci_year = typhoon_report.sub2_sdate[:4]
        acc.sub_group_id = str(typhoon_report.mem_num).strip()
        acc.school_name = typhoon_report.school_name
        acc.school_number = typhoon_report.school_num
#             if typhoon_report.organ_gubun2_name in ['초등학교','중학교','고등학교','대학'] :
#                 acc.school_type = typhoon_report.organ_gubun2_name
#             else:
#                 acc.school_type = '기타'
        acc.school_type = typhoon_report.organ_gubun2_name
        acc.school_address = typhoon_report.addr1
        acc.latitude = float(typhoon_report.lat)
        acc.longitude = float(typhoon_report.lng)
        acc.acci_occu_date = typhoon_report.sub2_sdate
        acc.establish_gubun_name = typhoon_report.establish_gubun_name
        # acc.damage_desc = item['']
        if typhoon_report.sub2_data1 != '':
            try :
                acc.precipitation = float(typhoon_report.sub2_data1)
            except :
                acc.precipitation = 0
        else:
            acc.precipitation = 0
        if typhoon_report.sub2_data3 != '':
            try:
                acc.temperature = float(typhoon_report.sub2_data3)
            except:
                acc.temperature = 0
        if typhoon_report.sub2_data4 != '':
            try :
                acc.wind_speed = float(typhoon_report.sub2_data4)
            except :
                acc.wind_speed = 0
        else :
            acc.wind_speed = 0
        acc.typhoon_name = typhoon_report.sub2_disaster_naming
        acc.total_cost = typhoon_report.sub6_total_payment_money

        acc.fk_typhoon_id = typhoon_report.fk_typhoon_id

        sc = db.session.query(Schools).filter(Schools.school_number == acc.school_number).first()
        if sc is not None:
            acc.fk_school_id = sc.id
        else:
            acc.fk_school_id = -1

        ar = db.session.query(Areas).filter(Areas.area_name.like("%" + typhoon_report.area_code_name + "%")).first()
        if ar is not None:
            acc.fk_area_id = ar.id
        else:
            acc.fk_area_id = -1

        group = db.session.query(SubGroups).filter(SubGroups.sub_group_id == acc.sub_group_id).first()
        if group is not None:
            acc.fk_sub_group_id = group.id
        else:
            acc.fk_sub_group_id = -1

        acc.s3_data_json =typhoon_report.s3_data_json
        DBManager.db.session.add(acc)
#     else :
#         print("school name :",typhoon_report.school_name ,",idx :",idx + 1)
#         object = {
#             "total_cost":typhoon_report.sub6_total_payment_money,
#             "s3_data_json":typhoon_report.s3_data_json
#         }
#         DBManager.db.session.query(Accidents).filter_by(acci_number=typhoon_report.disaster_num).update(object)

    DBManager.db.session.commit()

    return make_response(jsonify(''), 200)


@app.route('/api/monitor/v1/data_file_move', methods=['POST'])
def data_file_move_api():
    from shutil import copyfile
    typhoon_report_list = TB_TYPOONREPORT.query \
        .filter(TB_TYPOONREPORT.sdate >='2020-01-01') \
        .filter(TB_TYPOONREPORT.sdate <'2021-01-01') \
        .all()
    print("len :",len(typhoon_report_list))
    #print("Accident add start")
    SRC_FILE_PATH = 'D:\\project\\typhoon\\typhoon\data\\'
    DEST_FILE_PATH = 'D:\\project\\typhoon\\typhoon\\frontend_vue\\public\\data\\'
#     global skip
#     skip = True
    def find_file(name, path):
        for root, dirs, files in os.walk(path):
            if name in files:
                return os.path.join(root, name)

#         global skip
#         if name == '202004221511211.pdf':
#             skip = False
#         if skip == False :
#             for root, dirs, files in os.walk(path):
#                 if name in files:
#                     return os.path.join(root, name)
    def db_file_copy(db_file_name):
        if db_file_name is not None :
            file_name = db_file_name.split(';')[-1]
            if file_name is not None :
                find_name = find_file(file_name,SRC_FILE_PATH)
                print("find_name :",find_name)
                if find_name is not None :
                    copyfile(find_name,DEST_FILE_PATH + file_name)

    for idx,typhoon_report in enumerate(typhoon_report_list):
        print("================> ", idx)
        db_file_copy(typhoon_report.basefile1)
        db_file_copy(typhoon_report.image_path1)
        db_file_copy(typhoon_report.image_path2)
        db_file_copy(typhoon_report.image_path3)
        db_file_copy(typhoon_report.image_path4)
        db_file_copy(typhoon_report.image_path5)
        db_file_copy(typhoon_report.image_path6)
        db_file_copy(typhoon_report.image_path7)
        db_file_copy(typhoon_report.image_path8)
    return make_response(jsonify(''), 200)

@app.route('/api/monitor/v1/acci_image_path', methods=['GET'])
@token_required
def acci_image_path_api():
    parser = reqparse.RequestParser()
    parser.add_argument("acci_id", type=str, location='args', required=True)
    args = parser.parse_args()
    response = dict()
    acci_id = args['acci_id']
    print("acci_id:",acci_id)
    acci = Accidents.query.filter_by(id=acci_id).first()
    if acci is not None:
        typhoon = Typhoons.query.filter_by(id=acci.fk_typhoon_id).first()
        typhoon_name = typhoon.typhoon_name.split('(')[0]
        school_num = acci.school_number
        print("typhoon_name:",typhoon_name,",school_num:",school_num)
        typhoon_report = TB_TYPOONREPORT.query.filter_by(school_num=school_num).first()
        typhoon_report = TB_TYPOONREPORT.query.filter_by(school_num=school_num).filter(TB_TYPOONREPORT.sub2_disaster_naming.like("%" + str(typhoon_name) +"%")).first()
        if typhoon_report is not None :
#             print("disaster_num:",typhoon_report.disaster_num)
            if len(typhoon_report.image_file_path1) > 10 :
                first_idx = typhoon_report.image_file_path1.find('src=')
                end_idx = typhoon_report.image_file_path1.find('.png')
                image_file_path = typhoon_report.image_file_path1[first_idx+5:end_idx+4]
                print("image_file_path :",image_file_path)
                response['image_file_path1'] = image_file_path
    return make_response(jsonify(response), 200)

@app.route('/api/monitor/v1/report/report_acci_school', methods=['GET'])
@token_required
def report_acci_school_api():
    parser = reqparse.RequestParser()
    parser.add_argument("offset", type=int, location='args', required=True)
    parser.add_argument("limit", type=int, location='args')
    parser.add_argument("order_by", type=str, location='args')
    parser.add_argument("area", type=int, location='args')
    parser.add_argument("user_id", type=str, location='args')
    parser.add_argument("search_string", type=str, location='args')
    parser.add_argument("search_type", type=str, location='args')
    parser.add_argument("yearfrom", type=str, location='args')
    parser.add_argument("yearto", type=str, location='args')

    args = parser.parse_args()
    response = dict()

    offset = int(args['offset'])
    limit = int(args['limit'])
    order_by = args['offset']
    area = args['area']
    user_id = args['user_id']
    search_string = args['search_string']
    search_type = int(args['search_type'])
    yearfrom = datetime(int(args['yearfrom']), 1, 1, 00, 00, 00)
    yearto = datetime(int(args['yearto']), 12, 31, 23, 59, 59)

    print("input :",args)

    area_code_name = ''
    filter_and_group = []
    filter_or_group = []
    filter_and_group.append(TB_TYPOONREPORT.sub2_sdate >= yearfrom)
    filter_and_group.append(TB_TYPOONREPORT.sub2_sdate < yearto)
#     if area != None and search_type != 0:
    if area != 0 and area != None : #전체
        area_ = Areas.query.filter_by(id=area).first()
        print("area_.area_name :", area_.area_name)
        filter_and_group.append(TB_TYPOONREPORT.area_code_name == area_.area_name.strip())

    user = Users.query.filter_by(user_id=user_id).first()
    if user is not None and user.user_type == 5:
        filter_and_group.append(TB_TYPOONREPORT.mem_num == user_id)
    if user is not None and user.user_type == 6:
        filter_and_group.append(TB_TYPOONREPORT.school_num == user_id)
    if len(search_string) > 0 :
        if search_type == 0:
            print("search_type :",search_type,",search_string :",search_string)
            filter_or_group.append(TB_TYPOONREPORT.mem_name.like("%" + search_string + "%"))
            filter_or_group.append(TB_TYPOONREPORT.school_name.like("%" + search_string + "%"))
            filter_or_group.append(TB_TYPOONREPORT.sub2_disaster_naming.like("%" + search_string + "%"))
            filter_or_group.append(TB_TYPOONREPORT.disaster_num.like("%" + search_string + "%"))
        elif  search_type == 1:
            filter_and_group.append(TB_TYPOONREPORT.mem_name.like("%" + search_string + "%"))
        elif search_type == 2:
            filter_and_group.append(TB_TYPOONREPORT.school_name.like("%" + search_string + "%"))
        elif search_type == 3:
            filter_and_group.append(TB_TYPOONREPORT.sub2_disaster_naming.like("%" + search_string + "%"))
        elif search_type == 4:
            filter_and_group.append(TB_TYPOONREPORT.disaster_num.like("%" + search_string + "%"))

#     for filter in filter_and_group:
#         print("filter :",filter)
#     filter_and_group.append(db.func.substr(TB_TYPOONREPORT.sdate,1,4) == year_num)
#     if school_type != None:
#         filter_and_group.append(TB_TYPOONREPORT.organ_gubun2_name == school_type)

#
# #         print("area_.area_name :",area_.area_name)
#         if area_ is not None:
#             area_code_name = area_.area_name
#             filter_and_group.append(TB_TYPOONREPORT.area_code_name == area_code_name.strip())

    acci_schools = []
    total_count = TB_TYPOONREPORT.query.filter(or_(*filter_or_group)).filter(and_(*filter_and_group)).count()
    acci_school_list = TB_TYPOONREPORT.query.filter(or_(*filter_or_group)).filter(and_(*filter_and_group)).limit(limit).offset(offset)
    print("total_count :", total_count)
    for acci_school in acci_school_list:
        print("acci_school.school_name :",acci_school.school_name)
        sub2_disaster_naming = ''
        typhoon = Typhoons.query.filter_by(id=acci_school.fk_typhoon_id).first()
        if typhoon is not None :
            sub2_disaster_naming = get_typhoon_name_from_db(typhoon)
        else :
            sub2_disaster_naming = acci_school.sub2_disaster_naming
        acci_schools.append({
            "school_name":acci_school.school_name,
            "school_num":acci_school.school_num,
            "mem_name":acci_school.mem_name,
            "disaster_num":acci_school.disaster_num,
            "sub2_disaster_naming":sub2_disaster_naming,
            "building_info":getJsonGroup12(acci_school.s3_data_json),
#             "qurantee_info":get_school_info_from_json(acci_school.s3_data_json,'s3_2_group','s3_total_3','s3_2_group','s3_total_4'),
            "accessories":get_school_info_from_json_ext(acci_school,acci_school.s3_data_json,'s3_5_group','s3_total_7'),
            "article":get_school_info_from_json_ext(acci_school,acci_school.s3_data_json,'s3_5_group','s3_total_5'),
            "compensation":get_school_info_from_json_ext(acci_school,acci_school.s3_data_json,'s3_9_10','s3_total_9','s3_9_10','s3_total_10'),
            "payment_money":acci_school.sub6_total_payment_money
        })

#     # 태풍별 피해학교(막대), 태풍별 피해액수(막대)
#     tp_data = list()
#     query = db.session.query(Typhoons).filter(Typhoons.typhoon_number.like(str(year_num) + "%")).filter(Typhoons.typhoon_v_type != 1)
#
#     tps = query.all()
#     for tp in tps:
#         data = db.session.query(Accidents.id, db.func.sum(Accidents.total_cost).label("TOTAL_COST"),
#                                    db.func.count(Accidents.id).label("TOTAL_CN")).filter(Accidents.fk_typhoon_id == tp.id)
#
#         if school_type is not None:
#             data = data.filter(Accidents.school_type == school_type)
#         if area is not None:
#             data = data.filter(Accidents.fk_area_id == area)
#         data = data.first()
#
#         tp_data.append(dict(tp_name=tp.typhoon_name,
#                             acci_cost=str(data.TOTAL_COST) if data.TOTAL_COST is not None else "0",
#                             school_cnt=str(data.TOTAL_CN) if data.TOTAL_CN is not None else "0"))
#     response['tp_data'] = tp_data
#     response['area_data'] = None
#     # 지역별 피해학교(막대), 지역별 피해액수(막대)
#     if area is not None:
#         response['area_data'] = None
#     else:
#         region_data = list()
#
#         areas = db.session.query(Areas).all()
#         for area_ in areas:
#             data = db.session.query(Accidents.id, db.func.sum(Accidents.total_cost).label("TOTAL_COST"),
#                                     db.func.count(Accidents.id).label("TOTAL_CN")).filter(Accidents.fk_area_id == area_.id).first()
#             data = data.first()
#             region_data.append(dict(area_name=area_.area_name,
#                                 acci_cost=str(data.TOTAL_COST) if data.TOTAL_COST is not None else "0",
#                                 school_cnt=str(data.TOTAL_CN) if data.TOTAL_CN is not None else "0"))
#         response['area_data'] = region_data
    response = {
        "num_results":total_count,
        "objects":acci_schools
    }
    return make_response(jsonify(response), 200)


#
@app.route('/api/monitor/v1/damage_manage_download', methods=['get'])
def damage_manage_download_api():
#         let params = {
#           search_string:this.accident.search_text,
#           start_date:this.search.dialog.form.acci_start_date,
#           end_date:this.search.dialog.form.acci_end_date,
#           search_type_code:this.search.dialog.selected_search_type_code,
#           tab_item:this.selected_tab_item,
#           user_type:this.$session.getUserType()
#         }

    parser = reqparse.RequestParser()
    parser.add_argument("search_string", type=str, location='args', required=True)
    parser.add_argument("start_date", type=str, location='args', required=True)
    parser.add_argument("end_date", type=str, location='args', required=True)
    parser.add_argument("tab_item", type=str, location='args', required=True)
    parser.add_argument("search_type_code", type=str, location='args', required=True)
    parser.add_argument("user_type", type=str, location='args', required=True)
    parser.add_argument("user_id", type=str, location='args', required=True)
    args = parser.parse_args()
    search_string = args['search_string']
    start_date = args['start_date']
    end_date = args['end_date']
    tab_item = int(args['tab_item'])
    search_type_code = int(args['search_type_code'])
    user_type = int(args['user_type'])
    user_id = args['user_id']

    if tab_item == 0:
        templete_path = './download/excel/templete/damage_manage_xl_template0.xlsx'
        xl_range = 10
    else:
        templete_path = './download/excel/templete/damage_manage_xl_template1.xlsx'
        xl_range = 7

    #xl파일 형식지정
    center_align = Alignment(horizontal='center',
                            vertical='bottom',
                            text_rotation=0,
                            wrap_text=False,
                            shrink_to_fit=True,
                            indent=0)

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    filter_or_group = []
    filter_and_group = []
    if tab_item == 0:
        if search_string != '' :
            if search_type_code == 0 :
                filter_or_group.append(TB_TYPOONREPORT.school_name.like("%" + search_string + "%"))
                filter_or_group.append(TB_TYPOONREPORT.disaster_num.like("%" + search_string + "%"))
                filter_or_group.append(TB_TYPOONREPORT.mem_name.like("%" + search_string + "%"))
                filter_or_group.append(TB_TYPOONREPORT.school_num.like("%" + search_string + "%"))
                filter_or_group.append(TB_TYPOONREPORT.addr1.like("%" + search_string + "%"))
                filter_or_group.append(TB_TYPOONREPORT.sub2_disaster_naming.like("%" + search_string + "%"))
            elif search_type_code == 1 :
                filter_and_group.append(TB_TYPOONREPORT.mem_name.like("%" + search_string + "%"))
            elif search_type_code == 2 :
                filter_and_group.append(TB_TYPOONREPORT.school_name.like("%" + search_string + "%"))
            elif search_type_code == 3 :
                filter_and_group.append(TB_TYPOONREPORT.school_num.like("%" + search_string + "%"))

        filter_and_group.append(TB_TYPOONREPORT.sub2_sdate >= start_date)
        filter_and_group.append(TB_TYPOONREPORT.sub2_sdate <= end_date)
        if user_type == 5 :
            filter_and_group.append(TB_TYPOONREPORT.mem_num == user_id)
        elif user_type == 6 :
            filter_and_group.append(TB_TYPOONREPORT.school_num == user_id)
        damage_data = TB_TYPOONREPORT.query.filter(or_(*filter_or_group)).filter(and_(*filter_and_group)).order_by(TB_TYPOONREPORT.sub2_sdate.desc()).all()
    else:
        if search_string != '' :
            if search_type_code == 0 :
                filter_or_group.append(AccidentsPredict.school_name.like("%" + search_string + "%"))
                filter_or_group.append(AccidentsPredict.area_code_name.like("%" + search_string + "%"))
                filter_or_group.append(AccidentsPredict.mem_name.like("%" + search_string + "%"))
                filter_or_group.append(AccidentsPredict.establish_gubun_name.like("%" + search_string + "%"))
                filter_or_group.append(AccidentsPredict.organ_gubun2_name.like("%" + search_string + "%"))
            elif search_type_code == 1 :
                filter_and_group.append(AccidentsPredict.mem_name.like("%" + search_string + "%"))
            elif search_type_code == 2 :
                filter_and_group.append(AccidentsPredict.school_name.like("%" + search_string + "%"))
            elif search_type_code == 3 :
                filter_and_group.append(AccidentsPredict.sub2_disaster_naming.like("%" + search_string + "%"))

        if user_type == 5 :
            filter_and_group.append(AccidentsPredict.sub_group_num == user_id)
        elif user_type == 6 :
            filter_and_group.append(AccidentsPredict.sub_school_num == user_id)
#역순으로 됨
        damage_data = AccidentsPredict.query.filter(or_(*filter_or_group)).filter(and_(*filter_and_group)).order_by(AccidentsPredict.id.desc()).all()

    wb = openpyxl.load_workbook(templete_path)
    ws = wb['Sheet1']

    def getLevel(cost):
        cost_array = [10000000,20000000,50000000,100000000]
        if cost < cost_array[0] :
            return 1
        elif cost >= cost_array[0] and cost <= cost_array[1] :
            return 2
        elif cost >= cost_array[1] and cost <= cost_array[2] :
            return 3
        elif cost >= cost_array[2] and cost < cost_array[3]:
            return 4
        else :
            return 5

    def get_cost_level(cost):
      cost_array = [10000000, 20000000, 50000000, 100000000]
      if cost < cost_array[0]:
        return '규모 1(관심)'
      elif cost >= cost_array[0] and cost <= cost_array[1]:
        return '규모 2(주의)'
      elif cost >= cost_array[1] and cost <= cost_array[2]:
        return '규모 3(경계)'
      elif cost >= cost_array[2] and cost < cost_array[3]:
        return '규모 4(위험)'
      else :
        return '규모 5(심각)'
    
    cur=0
    for rows in damage_data:
        typhoon_name = rows.sub2_disaster_naming
        typhoon = Typhoons.query.filter_by(id=rows.fk_typhoon_id).first()
        if typhoon is not None :
            typhoon_name = typhoon.typhoon_name
        if tab_item == 0:
            line_tuple = (
                rows.disaster_num,
                rows.mem_name,
                rows.school_name,
                rows.area_code_name,
                rows.establish_gubun_name,
                rows.organ_gubun2_name,
                typhoon_name,
                rows.sub2_sdate,
                rows.sub2_damage_detail,
                f'{rows.total_payment_money:,}' if rows.total_payment_money is not None else 0
            )
        else:
            line_tuple = (
                rows.mem_name,
                rows.school_name,
                rows.area_code_name,
                rows.establish_gubun_name,
                rows.organ_gubun2_name,
                rows.sub2_disaster_naming,
                get_cost_level(rows.total_payment_money) if rows.total_payment_money is not None else 0
            )
        for i in range(xl_range):
            ws.cell(row=cur+1, column=i+1).border = thin_border
            ws.cell(row=cur+1, column=i+1).alignment = center_align
        cur+=1
        print(line_tuple)
        ws.append(line_tuple)

    for k in range(3):
        for i in range(xl_range):
            ws.cell(row=cur+1, column=i+1).border = thin_border
            ws.cell(row=cur+1, column=i+1).alignment = center_align
        cur+=1
    global download_index
    download_index += 1
    save_file_name = 'download_xl_file_{0}.xlsx'.format(download_index)
    save_file_path = './download/excel/' + save_file_name
    wb.save(save_file_path)
    res = {
        "filename":save_file_name
    }

#     print("type:",tab_item)
#     print(type(tab_item))
#     print("templete:",templete_path)
#     print('xl_range:',xl_range)
    return make_response(jsonify(res), 200)

@app.route('/api/monitor/v1/damage_detail_download', methods=['get'])
def damage_detail_download_api():

    parser = reqparse.RequestParser()
    parser.add_argument("search_string", type=str, location='args', required=True)
    parser.add_argument("area", type=int, location='args')
    parser.add_argument("user_id", type=str, location='args', required=True)
    parser.add_argument("tab_item", type=str, location='args', required=True)
    parser.add_argument("search_type", type=str, location='args', required=True)
    parser.add_argument("yearfrom", type=str, location='args', required=True)
    parser.add_argument("yearto", type=str, location='args', required=True)

    args = parser.parse_args()
    search_string = args['search_string']
    area = args['area']
    user_id = args['user_id']
    tab_item = int(args['tab_item'])
    search_type = int(args['search_type'])
    yearfrom = datetime(int(args['yearfrom']), 1, 1, 00, 00, 00)
    yearto = datetime(int(args['yearto']), 12, 31, 23, 59, 59)


    templete_path = './download/excel/templete/damage_detail_xl_template.xlsx'
    xl_range = 9

    #xl파일 형식지정
    center_align = Alignment(horizontal='center',
                            vertical='bottom',
                            text_rotation=0,
                            wrap_text=False,
                            shrink_to_fit=True,
                            indent=0)

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    area_code_name = ''
    filter_and_group = []
    filter_or_group = []
    filter_and_group.append(TB_TYPOONREPORT.sub2_sdate >= yearfrom)
    filter_and_group.append(TB_TYPOONREPORT.sub2_sdate < yearto)

    if area != 0 and area != None : #전체
        area_ = Areas.query.filter_by(id=area).first()
        filter_and_group.append(TB_TYPOONREPORT.area_code_name == area_.area_name.strip())

    user = Users.query.filter_by(user_id=user_id).first()
    if user is not None and user.user_type == 5:
        filter_and_group.append(TB_TYPOONREPORT.mem_num == user_id)
    if user is not None and user.user_type == 6:
        filter_and_group.append(TB_TYPOONREPORT.school_num == user_id)

    if len(search_string) > 0 :
        if search_type == 0:
            print("search_type :",search_type,",search_string :",search_string)
            filter_or_group = []
            filter_or_group.append(TB_TYPOONREPORT.mem_name.like("%" + search_string + "%"))
            filter_or_group.append(TB_TYPOONREPORT.school_name.like("%" + search_string + "%"))
            filter_or_group.append(TB_TYPOONREPORT.sub2_disaster_naming.like("%" + search_string + "%"))
            filter_or_group.append(TB_TYPOONREPORT.disaster_num.like("%" + search_string + "%"))
        elif search_type == 1:
            filter_and_group.append(TB_TYPOONREPORT.mem_name.like("%" + search_string + "%"))
        elif search_type == 2:
            filter_and_group.append(TB_TYPOONREPORT.school_name.like("%" + search_string + "%"))
        elif search_type == 3:
            filter_and_group.append(TB_TYPOONREPORT.sub2_disaster_naming.like("%" + search_string + "%"))
        elif search_type == 4:
            filter_and_group.append(TB_TYPOONREPORT.disaster_num.like("%" + search_string + "%"))

    acci_schools = []
    damage_data = TB_TYPOONREPORT.query.filter(or_(*filter_or_group)).filter(and_(*filter_and_group)).all()

    wb = openpyxl.load_workbook(templete_path)
    ws = wb['Sheet1']
    cur=0

    def get_parse_set(data) :
        count = data['count']
        cost = data['cost'] if data['cost'] is not None else 0
        # print("data count:",data['count'])
        # print("data cost:",data['cost'])
        return str(count) + '건(' + f'{int(cost):,}' + ')'
    
    for rows in damage_data:
        typhoon_name = ''
        typhoon = Typhoons.query.filter_by(id=rows.fk_typhoon_id).first()
        if typhoon is not None :
            typhoon_name = typhoon.typhoon_name
        line_tuple = (
            rows.disaster_num,
            rows.mem_name,
            rows.school_name,
            typhoon_name,
            get_parse_set(getJsonGroup12(rows.s3_data_json)),
#             get_parse_set(get_school_info_from_json(rows.s3_data_json,'s3_2_group','s3_total_3','s3_2_group','s3_total_4')),
            get_parse_set(get_school_info_from_json_ext(rows,rows.s3_data_json,'s3_5_group','s3_total_7')),
            get_parse_set(get_school_info_from_json_ext(rows,rows.s3_data_json,'s3_5_group','s3_total_5')),
            get_parse_set(get_school_info_from_json_ext(rows,rows.s3_data_json,'s3_9_10','s3_total_9','s3_9_10','s3_total_10')),
            f'{rows.sub6_total_payment_money:,}' if rows.sub6_total_payment_money is not None else 0
        )
#         print("line_tuple :",line_tuple)

        for i in range(xl_range):
            ws.cell(row=cur+1, column=i+1).border = thin_border
            ws.cell(row=cur+1, column=i+1).alignment = center_align
        cur+=1
#         print(line_tuple)
        ws.append(line_tuple)

    for k in range(3):
        for i in range(xl_range):
            ws.cell(row=cur+1, column=i+1).border = thin_border
            ws.cell(row=cur+1, column=i+1).alignment = center_align
        cur+=1
    global download_index
    download_index += 1
    save_file_name = 'download_detail_file_{0}.xlsx'.format(download_index)
    save_file_path = './download/excel/' + save_file_name
    wb.save(save_file_path)
    res = {
        "filename":save_file_name
    }

    return make_response(jsonify(res), 200)

@app.route('/api/monitor/v1/user_password_init', methods=['POST'])
def user_password_init_api():
    data = json.loads(request.data)
    result = ''
    print("user_password_init_api start")
    user_id = data['user_id']
    user_name = data['user_name']
    telephone = data['telephone']
    print("data : ", data)
    def genPass():
      alphabet = "abcdefghijklmnopqrstuvwxyz0123456789"
      password = ""

      for i in range(6):
          index = random.randrange(len(alphabet))
          password = password + alphabet[index]
      return password
    smtp = Settings.query.filter_by(id=1).first()
    if smtp is not None :
        user = Users.query.filter_by(user_id=user_id).filter_by(user_name=user_name).filter_by(telephone=telephone).first()
        if user is not None :
            print("1")
            alphabet = "abcdefghijklmnopqrstuvwxyz0123456789"
            passwd = genPass() + 'QW!@'
            message = '''<pre>
        교육시설 재난피해(태풍) 예측 시스템(KO-Dis)의 비밀번호가 초기화되었습니다.
        아래 비밀번호로 로그인하여주시기 바랍니다.

        초기화 비밀번호 : {0}

        ※ 로그인 후 내 정보 수정을 통해 비밀번호를 수정하여 주시기 바랍니다.

        감사합니다.
            <pre>'''.format(str(passwd))
            print("message :",message)
            server = smtplib.SMTP_SSL(smtp.smtp_server, smtp.smtp_port)
            server.login(smtp.smtp_id, smtp.smtp_pass)


            msg = MIMEMultipart('alternative')
            msg['From'] = "%s <%s>" % ("", smtp.smtp_sender)
            msg['To'] = user.email
            msg['Subject'] = "[한국교육시설안전원] 교육시설 재난피해(태풍) 예측 시스템 비밀번호 초기화 메일"
            msg.attach(MIMEText(message, 'html', 'utf-8'))  # 내용 인코딩
            server.sendmail(smtp.smtp_sender, user.email, msg.as_string())
            print("msg :",msg)
            Users.query.filter(Users.id == user.id).update({'user_pw': password_encoder_512(passwd)})
            db.session.commit()
            print("complete passwd :", password_encoder_512(passwd))
            result = {'status': True}
        else :
            print("2")
            result = {'status': False}
    return make_response(jsonify(result), 200)

@app.route('/api/v1/member_address_update', methods=['GET'])
def member_address_update_api():
    print('member_address_update_api ==> run')
    file_path = 'C:\\d\\project\\typhoon\\data\\온라인공제관리시스템_회원 주소록_210927 기준.xlsx'
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    for cur,file_data in enumerate(ws):
        if cur < 4:
            continue
        cat = SubGroups()
        file_data = [None if data.value==None else data.value if data.value[-1]!='\xa0' else data.value[:-1] for data in file_data]
#         fk_sub_group_id = -1
#         fk_area_id = -1

        grp = db.session.query(SubGroups).filter(SubGroups.sub_group_id.like("%" + file_data[0] + "%")).first()
        if grp is None:
            cat.sub_group_id = file_data[0]
            cat.sub_group_name = file_data[1]
            cat.sub_group_type = file_data[2]
            cat.group_location = file_data[3]
            cat.user_name = file_data[4]
            cat.user_email = file_data[5]
            cat.user_tel = file_data[6]
            cat.user_mobile = file_data[7]
            cat.sub_group_post = file_data[8]
            cat.sub_group_address = file_data[9]
            cat.created_date = datetime.now()
            cat.modify_date = datetime.now()
            db.session.add(cat)
            print('add :',cur)
            db.session.commit()

    print('api end')
    return make_response(jsonify({}),200)

@app.route('/api/v1/school_address_update', methods=['GET'])
def school_address_update_api():
    print('school_address_update_api ==> run')
    file_path = 'C:\\d\\project\\typhoon\\data\\온라인공제관리시스템_학교 주소록_210927 기준.xlsx'
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
#     db_data = SubSchools.query.all()
#     db_ids = [db_datas.sub_school_id for db_datas in db_data]

    for cur,file_data in enumerate(ws):
        if cur < 4:
            continue
        cat = SubSchools()
        file_data = [None if data.value==None else data.value if data.value[-1]!='\xa0' else data.value[:-1] for data in file_data]

        sch = db.session.query(SubSchools).filter(SubSchools.sub_school_id.like("%" + file_data[0] + "%")).first()
        if sch is not None:
            obj = {
                'sub_school_type':file_data[6],
                'sub_school_tel':file_data[8],
                'sub_school_address':file_data[10]
            }
            db.session.query(SubSchools).filter_by(sub_school_id=file_data[0]).update(obj)
            print('update :',cur)
            db.session.commit()
        else:
            cat.sub_school_id = file_data[0]
            cat.sub_school_name = file_data[1]
            cat.sub_group_name = file_data[2]
            cat.sub_school_email = file_data[3]
            cat.sub_school_sub = file_data[4]
            cat.sub_school_loc = file_data[5]
            cat.sub_school_type = file_data[6]
            cat.sub_school_alive_type = file_data[7]
            cat.sub_school_tel = file_data[8]
            cat.sub_school_mobile = file_data[9]
            cat.sub_school_address = file_data[10]
            grp_id = db.session.query(SubGroups).filter(SubGroups.sub_group_name.like("%" + file_data[2] + "%")).first()
            cat.fk_sub_group_id = grp_id.id
            cat.created_date = datetime.now()
            cat.modify_date = datetime.now()
            ar_id = db.session.query(Areas).filter(Areas.area_name.like("%" + file_data[5] + "%")).first()
            cat.fk_area_id = ar_id.id
            db.session.add(cat)
            print('add :',cur)
            db.session.commit()

    print('api end')
    return make_response(jsonify({}),200)

@app.route('/api/v1/school_tudes_update', methods=['GET'])
def school_tudes_update_api():
    print('school_info_update_api ==> run')
#     file_path = 'C:\\d\\project\\typhoon\\data\\교육연구시설 건물별 건축년도 및 관리현황_20201019_1 (1).xlsx'
#     wb = openpyxl.load_workbook(file_path)
#     ws1 = wb['공제회데이터']
#     ws2 = wb['공제회데이터 2']
    db_datas = SubSchools.query.all()
    print('data load')

    for cur,db_data in enumerate(db_datas):
        #1경도위도
        if db_data.latitude == None and db_data.sub_school_address is not None:
#             address = db_data.sub_school_address
            url = "https://dapi.kakao.com/v2/local/search/address.json?&query=" + db_data.sub_school_address
            res = requests.get(urlparse(url).geturl(),headers={"Authorization": "KakaoAK 34e91abb80d18b393750c226da9649a1"})
            json_obj = res.json()
            for document in json_obj['documents']:
                val = [document['y'], document['x']]
                latitude = str("%0.2f" % float(document['y']))
                longitude = str("%0.2f" % float(document['x']))
#                 print("latitude :",latitude,",logitude:",longitude)
                obj = {
                    "latitude":latitude,
                    "longitude":longitude
                }
                db.session.query(SubSchools).filter_by(id=db_data.id).update(obj)
                db.session.commit()
                print('cur :',cur)
                break
        #2설립연도
#         if db_data.sub_school_establish == None:
    print('api end')
    return make_response(jsonify({}),200)

import re
@app.route('/api/v1/school_establish_update', methods=['GET'])
def school_establish_update_api():
    def roof1(ws):
        print('roof1 run')
        cnt = 0
        for cur,ws_data in enumerate(ws):
            ws_data_list = [data.value for data in ws_data]

            if ws_data_list[0] is not None:
#                 sch_estas = db.session.query(SubSchools).filter(SubSchools.sub_school_name.like('%'+ws_data_list[0]+'%')).all()
#                 sch_esta = [se for se in sch_estas if len(se.sub_school_name)==len(ws_data_list[0])+1]
                sch_esta = db.session.query(SubSchools).filter(SubSchools.sub_school_establish.is_(None)).filter(SubSchools.sub_school_name.like('%'+ws_data_list[0]+'%')).all()
                esta = re.sub(r'[^0-9]', '', ws_data_list[1])
                esta = None if esta is '' else esta
                if esta is None:
                    continue
                if len(sch_esta)>1:
                    for sch_esta1 in sch_esta:
                        obj = {'sub_school_establish':esta}
                        db.session.query(SubSchools).filter_by(id=sch_esta1.id).update(obj)
                        db.session.commit()
                        cnt+=1
                        print(sch_esta1,':',obj)
                        print('1-id :'+sch_esta1.sub_school_name+'/  cur :',cur,' /  cnt :',cnt,'\n')
                else:
                    if len(sch_esta)==0:
                        continue
                    sch_esta1 = sch_esta[0]
                    print('sch_esta1:',sch_esta1)
                    obj = {'sub_school_establish':esta}
                    db.session.query(SubSchools).filter_by(id=sch_esta1.id).update(obj)
                    db.session.commit()
                    cnt+=1
                    print(sch_esta1,':',obj)
                    print('2-id :',sch_esta1.sub_school_name,' /  cur :',cur,' /  cnt :',cnt,'\n')

    print('school_info_update_api ==> run')
    file_path = 'C:\\d\\project\\typhoon\\data\\교육연구시설 건물별 건축년도 및 관리현황_20201019_1 (1).xlsx'
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws1 = wb['공제회데이터'].rows
    ws2 = wb['공제회데이터 2'].rows
    print('data load')

    roof1(ws1)
    roof1(ws2)

    print('api end')
    return make_response(jsonify({}),200)